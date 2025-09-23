import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.drawing.image import Image

logo_works = r'Y:\Power BI\Scritps\python\Projetos\logo_works.jpg'
logo_pressseg = r'Y:\Power BI\Scritps\python\Projetos\logo_pressseg.png'

def gerar_relatorio():
    try:
        mes = combo_mes.get()
        ano = entry_ano.get()

        if not mes or not ano.isdigit():
            messagebox.showerror("Erro", "Informe corretamente o mês e o ano!")
            return

        arquivo_excel = filedialog.askopenfilename(
            title="Selecione a base de dados",
            filetypes=[("Planilhas", "*.xlsx *.xls *.ods *.csv")]
        )
        if not arquivo_excel:
            return

        ext = os.path.splitext(arquivo_excel)[1].lower()
        if ext == ".xlsx":
            df = pd.read_excel(arquivo_excel, engine="openpyxl")
        elif ext == ".xls":
            df = pd.read_excel(arquivo_excel, engine="xlrd")
        elif ext == ".csv":
            df = pd.read_csv(arquivo_excel, sep=";", encoding="latin1")
        else:
            messagebox.showerror("Erro", "Formato de arquivo não suportado!")
            return

        df.columns = df.columns.str.lower().str.strip()

        coluna_posto = "posto"
        coluna_re = "re"
        coluna_nome = "nome"
        coluna_funcao = "desc_cargo"
        coluna_data = "data"
        coluna_hora = "hor_inicio"
        coluna_csituacao = "csituacao"
        coluna_csithoje = "csithoje"

        df[coluna_data] = pd.to_datetime(df[coluna_data], dayfirst=True, errors="coerce")
        df["dia"] = df[coluna_data].dt.day

        # Filtros
        df = df[df[coluna_csituacao].isin([10, 11])]
        df = df[~df[coluna_csithoje].isin([2, 13, 14])]

        # Definir turno
        def definir_turno(hora_inicio):
            try:
                if isinstance(hora_inicio, str) and ":" in hora_inicio:
                    h, m = map(int, hora_inicio.split(":"))
                    h = h * 100 + m
                elif pd.notna(hora_inicio):
                    h = int(hora_inicio)
                else:
                    return "INDEFINIDO"
                return "NOTURNO" if h >= 1200 else "DIURNO"
            except:
                return "INDEFINIDO"

        df["turno"] = df[coluna_hora].apply(definir_turno)

        # Nome da aba
        def limpar_nome_aba(nome):
            nome_limpo = re.sub(r'[:\\/*?\[\]]', '_', str(nome))
            return nome_limpo[:31]

        # Saída
        saida = filedialog.asksaveasfilename(
            title="Salvar Relatório",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="relatorio_por_posto.xlsx"
        )
        if not saida:
            return

        # Geração do Excel
        with pd.ExcelWriter(saida, engine="openpyxl") as writer:
            abas_criadas = 0
            for posto, bloco_posto in df.groupby(coluna_posto):
                if bloco_posto.empty:
                    continue

                resultado = (
                    bloco_posto.groupby([coluna_re, coluna_nome, coluna_funcao, "turno"])
                    [coluna_data]
                    .apply(lambda x: ",".join(str(d.day) for d in sorted(x.dropna().unique())))
                    .reset_index()
                    .rename(columns={
                        coluna_re: "RE",
                        coluna_nome: "NOME",
                        coluna_funcao: "CARGO",
                        "turno": "TURNO",
                        coluna_data: "DIAS TRABALHADOS"
                    })
                )

                # Quebra entre diurno e noturno
                diurnos = resultado[resultado["turno"] == "DIURNO"]
                noturnos = resultado[resultado["turno"] == "NOTURNO"]
                linha_vazia = pd.DataFrame([[""] * len(resultado.columns)], columns=resultado.columns)
                resultado_final = pd.concat([diurnos, linha_vazia, noturnos], ignore_index=True)

                nome_aba = "Indefinido" if pd.isna(posto) else limpar_nome_aba(posto)
                resultado_final.to_excel(writer, sheet_name=nome_aba, index=False, startrow=14, startcol=0)
                abas_criadas += 1

            if abas_criadas == 0:
                resumo = pd.DataFrame({"Mensagem": ["Nenhum dado válido encontrado para gerar o relatório."]})
                resumo.to_excel(writer, sheet_name="Resumo", index=False)

        # --- Formatação ---
        wb = load_workbook(saida)
        fonte = Font(name="Arial", size=7)
        fonte_cnpj = Font(name="Arial", size=6)
        fonte_negrito = Font(name="Arial", size=8, bold=True)
        fonte_titulo = Font(name="Arial", size=10, bold=True)
        alinhamento_centro = Alignment(horizontal="center", vertical="center")
        borda = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        for aba in wb.sheetnames:
            ws = wb[aba]

            def aplicar_borda(range_str):
                start_cell, end_cell = range_str.split(":")
                start_col, start_row = ws[start_cell].column, ws[start_cell].row
                end_col, end_row = ws[end_cell].column, ws[end_cell].row
                for row in ws.iter_rows(min_row=start_row, max_row=end_row,
                                        min_col=start_col, max_col=end_col):
                    for cell in row:
                        cell.border = borda

            # Dados do cabeçalho
            contrato = entry_contrato.get() or "NÃO INFORMADO"
            pe = entry_pe.get() or "NÃO INFORMADO"
            processo = entry_processo.get() or "NÃO INFORMADO"

            ultima_coluna = "E"

            # Logo + Cabeçalho
            if combo_Empresa.get() == 'Works':
                if os.path.exists(logo_works):
                    img = Image(logo_works)
                    img.width, img.height = 134, 104
                    ws.add_image(img, "C1")

                cabecalhos = [
                    (f"A7:{ultima_coluna}7", "Endereço: R. Conselheiro Ribas, 297 - Vila Anastácio, São Paulo - SP, 05093-060", fonte),
                    (f"A8:{ultima_coluna}8", "CNPJ: 56.419.492/0001-09     Telefone: (11) 4563-9017", fonte),
                    (f"A9:{ultima_coluna}9", f"CONTRATO Nº {contrato} - P.E. {pe}     Processo Administrativo {processo}", fonte_cnpj)
                ]
            else:  # Pressseg
                if os.path.exists(logo_pressseg):
                    img = Image(logo_pressseg)
                    img.width, img.height = 251, 85
                    ws.add_image(img, "C1")

                cabecalhos = [
                    (f"A7:{ultima_coluna}7", "Endereço: R. Bernardo Guimarães, 210 - Vila Anastácio, São Paulo - SP, 05092-030", fonte),
                    (f"A8:{ultima_coluna}8", "CNPJ: 08818229/0001-40     Telefone: (11) 2507-2170", fonte),
                    (f"A9:{ultima_coluna}9", f"CONTRATO Nº {contrato} - P.E. {pe}     Processo Administrativo {processo}", fonte_cnpj)
                ]

            for range_str, texto, fnt in cabecalhos:
                start_cell = range_str.split(":")[0]
                ws.merge_cells(range_str)
                ws[start_cell].value = texto
                ws[start_cell].font = fnt
                ws[start_cell].alignment = alinhamento_centro
                aplicar_borda(range_str)

            # Nome do posto
            ws.merge_cells("A10:E10")
            ws["A10"].value = aba
            ws["A10"].font = fonte_titulo
            ws["A10"].alignment = alinhamento_centro
            aplicar_borda("A10:E10")

            # Mês referência
            ws.merge_cells("A11:E11")
            ws["A11"].value = f"Mês de Referência: 01 a 31 de {mes} de {ano}"
            ws["A11"].font = fonte_negrito
            ws["A11"].alignment = alinhamento_centro
            aplicar_borda("A11:E11")

            # Cabeçalho tabela (linha 15)
            for col in range(1, 6):
                cell = ws.cell(row=15, column=col)
                if cell.value:
                    cell.value = str(cell.value).lower()
                cell.font = fonte_negrito
                cell.alignment = alinhamento_centro
                cell.border = borda

            # Dados tabela
            for row in ws.iter_rows(min_row=16, max_row=ws.max_row, min_col=1, max_col=5):
                for cell in row:
                    if cell.value is not None:
                        cell.font = fonte
                        cell.alignment = alinhamento_centro
                        cell.border = borda

            # Rodapé
            ultima_linha = ws.max_row + 2
            meses = [
                "janeiro","fevereiro","março","abril","maio","junho",
                "julho","agosto","setembro","outubro","novembro","dezembro"
            ]
            try:
                indice_mes = meses.index(mes)
                mes_rodape = meses[(indice_mes + 1) % 12]
            except ValueError:
                mes_rodape = mes

            ws.merge_cells(start_row=ultima_linha, start_column=1, end_row=ultima_linha, end_column=5)
            ws.cell(row=ultima_linha, column=1).value = f"São Paulo, 01 de {mes_rodape} de {ano}"
            ws.cell(row=ultima_linha, column=1).font = fonte_negrito
            ws.cell(row=ultima_linha, column=1).alignment = alinhamento_centro

            # Assinaturas
            linha_assinatura = ultima_linha + 3
            ws.merge_cells(start_row=linha_assinatura, start_column=1, end_row=linha_assinatura, end_column=2)
            ws.cell(row=linha_assinatura, column=1).value = "________________________"
            ws.cell(row=linha_assinatura, column=1).alignment = Alignment(horizontal="left", vertical="center")

            ws.merge_cells(start_row=linha_assinatura, start_column=4, end_row=linha_assinatura, end_column=5)
            ws.cell(row=linha_assinatura, column=4).value = "________________________"
            ws.cell(row=linha_assinatura, column=4).alignment = Alignment(horizontal="left", vertical="center")

            ws.cell(row=linha_assinatura + 1, column=1).value = "Fiscal de Contrato"
            ws.cell(row=linha_assinatura + 1, column=1).font = fonte_negrito
            ws.cell(row=linha_assinatura + 1, column=1).alignment = Alignment(horizontal="left", vertical="center")

            ws.cell(row=linha_assinatura + 1, column=4).value = "Supervisão"
            ws.cell(row=linha_assinatura + 1, column=4).font = fonte_negrito
            ws.cell(row=linha_assinatura + 1, column=4).alignment = Alignment(horizontal="left", vertical="center")

        wb.save(saida)
        messagebox.showinfo("Sucesso", f"Relatório salvo em:\n{saida}")

    except Exception as e:
        messagebox.showerror("Erro", str(e))

# --- INTERFACE ---
root = tk.Tk()
root.title("📊 Gerador de Relatório por Posto e Turno")
root.geometry("500x350")
root.configure(bg="#f4f4f4")

style = ttk.Style()
style.theme_use("clam")
style.configure("TButton", font=("Segoe UI", 12, "bold"), padding=10,
                background="#4CAF50", foreground="white")
style.map("TButton", background=[("active", "#45a049")])

frame = ttk.Frame(root, padding=20)
frame.pack(expand=True)

titulo = tk.Label(frame, text="Gerador de Relatório por Posto e Turno",
                  font=("Segoe UI", 16, "bold"), bg="#f4f4f4", fg="#333")
titulo.pack(pady=10)

frame_data = ttk.Frame(frame)
frame_data.pack(pady=5)

tk.Label(frame_data, text="Empresa", font=("Segoe UI", 11), background="#f4f4f4").grid(row=0, column=0, padx=5)
combo_Empresa = ttk.Combobox(frame_data, values=["Works", "Pressseg"], width=12)
combo_Empresa.grid(row=0, column=1, padx=5)

tk.Label(frame_data, text="Mês:", font=("Segoe UI", 11), background="#f4f4f4").grid(row=0, column=2, padx=5)
combo_mes = ttk.Combobox(frame_data, values=[
    "janeiro", "fevereiro", "março", "abril", "maio", "junho",
    "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"
], width=12)
combo_mes.grid(row=0, column=3, padx=5)

tk.Label(frame_data, text="Ano:", font=("Segoe UI", 11), background="#f4f4f4").grid(row=0, column=5, padx=5)
entry_ano = tk.Entry(frame_data, width=8)
entry_ano.grid(row=0, column=6, padx=5)

# --- CAMPOS CONTRATO / P.E. / PROCESSO ---
frame_contrato = ttk.Frame(frame)
frame_contrato.pack(pady=5)

tk.Label(frame_contrato, text="Contrato:", font=("Segoe UI", 11), background="#f4f4f4").grid(row=0, column=0, padx=5)
entry_contrato = tk.Entry(frame_contrato, width=20)
entry_contrato.grid(row=0, column=1, padx=5)

tk.Label(frame_contrato, text="P.E.:", font=("Segoe UI", 11), background="#f4f4f4").grid(row=1, column=0, padx=5)
entry_pe = tk.Entry(frame_contrato, width=20)
entry_pe.grid(row=1, column=1, padx=5)

tk.Label(frame_contrato, text="Processo Adm.:", font=("Segoe UI", 11), background="#f4f4f4").grid(row=2, column=0, padx=5)
entry_processo = tk.Entry(frame_contrato, width=25)
entry_processo.grid(row=2, column=1, padx=5, columnspan=3)

btn = ttk.Button(frame, text="📂 Selecionar e Gerar Relatório", command=gerar_relatorio)
btn.pack(pady=20)

root.mainloop()
