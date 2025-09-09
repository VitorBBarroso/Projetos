import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.drawing.image import Image

logo = r'Y:\Power BI\Scritps\python\Projetos\logo_works.jpg'

def gerar_relatorio():
    try:
        mes = combo_mes.get()
        ano = entry_ano.get()
        
        if not mes or not ano.isdigit():
            messagebox.showerror("Erro", "Informe corretamente o mês e o ano!")
            return

        arquivo_excel = filedialog.askopenfilename(
            title="Selecione a base de dados",
            filetypes=[("Planilhas", "*.xlsx *.xls *.ods *.csv")]
        )
        if not arquivo_excel:
            return

        ext = os.path.splitext(arquivo_excel)[1].lower()
        if ext == ".xlsx":
            df = pd.read_excel(arquivo_excel, engine="openpyxl")
        elif ext == ".xls":
            df = pd.read_excel(arquivo_excel, engine="xlrd")
        elif ext == ".csv":
            df = pd.read_csv(arquivo_excel, sep=";", encoding="latin1")
        else:
            messagebox.showerror("Erro", "Formato de arquivo não suportado!")
            return

        df.columns = df.columns.str.lower().str.strip()

        coluna_posto = "posto"
        coluna_re = "re"
        coluna_nome = "nome"
        coluna_funcao = "desc_cargo"

        # --- Função auxiliar para limpar nomes de abas ---
        def limpar_nome_aba(nome):
            nome_limpo = re.sub(r'[:\\/*?\[\]]', '_', str(nome))
            return nome_limpo[:31]

        # --- Arquivo de saída ---
        saida = filedialog.asksaveasfilename(
            title="Salvar Relatório",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="relatorio_por_posto.xlsx"
        )
        if not saida:
            return

        with pd.ExcelWriter(saida, engine="openpyxl") as writer:
            abas_criadas = 0
            for posto, bloco_posto in df.groupby(coluna_posto):
                if bloco_posto.empty:
                    continue

                # Contar dias trabalhados (só para filtro)
                resultado = (
                    bloco_posto.groupby([coluna_re, coluna_nome, coluna_funcao])
                    .size()
                    .reset_index(name="dias_trabalhados")
                )

                # Filtrar apenas quem trabalhou >= 15 dias
                resultado = resultado[resultado["dias_trabalhados"] >= 15]

                if not resultado.empty:
                    # Remover a coluna "dias_trabalhados" (apenas filtro)
                    resultado = resultado.drop(columns=["dias_trabalhados"])

                    # Colunas em maiúsculo
                    resultado.columns = [col.upper() for col in resultado.columns]

                    nome_aba = "Indefinido" if pd.isna(posto) else limpar_nome_aba(posto)

                    # Escrever com cabeçalho na linha 15 (startrow=14 -> header ficará em linha 15)
                    resultado.to_excel(
                        writer,
                        sheet_name=nome_aba,
                        index=False,
                        startrow=14,   # header ficará na linha 15, dados a partir da 16
                        startcol=0,
                        header=True
                    )
                    abas_criadas += 1

            if abas_criadas == 0:
                resumo = pd.DataFrame({"Mensagem": ["Nenhum dado válido encontrado para gerar o relatório."]})
                resumo.to_excel(writer, sheet_name="Resumo", index=False)

        # --- Formatação Excel ---
        wb = load_workbook(saida)
        fonte = Font(name="Arial", size=7)
        fonte_cnpj = Font(name="Arial", size=6)
        fonte_negrito = Font(name="Arial", size=8, bold=True)
        fonte_titulo = Font(name="Arial", size=10, bold=True)
        alinhamento_centro = Alignment(horizontal="center", vertical="center")
        borda = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        for aba in wb.sheetnames:
            ws = wb[aba]

            # Logo centralizado
            if os.path.exists(logo):
                try:
                    img = Image(logo)
                    img.width, img.height = 134, 104
                    ws.add_image(img, "B1")
                except Exception:
                    pass  # não quebra se imagem tiver problema

            # Função para aplicar borda em células mescladas
            def aplicar_borda(range_str):
                start_cell, end_cell = range_str.split(":")
                start_col, start_row = ws[start_cell].column, ws[start_cell].row
                end_col, end_row = ws[end_cell].column, ws[end_cell].row
                for row in ws.iter_rows(min_row=start_row, max_row=end_row,
                                        min_col=start_col, max_col=end_col):
                    for cell in row:
                        cell.border = borda

            # --- pega os valores digitados ---
            contrato = entry_contrato.get() or "NÃO INFORMADO"
            pe = entry_pe.get() or "NÃO INFORMADO"
            processo = entry_processo.get() or "NÃO INFORMADO"

            # Cabeçalhos fixos
            cabecalhos = [
                ("A7:C7", "Endereço: R. Conselheiro Ribas, 297 - Vila Anastácio, São Paulo - SP, 05093-060", fonte),
                ("A8:C8", "CNPJ: 56.419.492/0001-09     Telefone: (11) 4563-9017", fonte),
                ("A9:C9", f"CONTRATO Nº {contrato} - P.E. {pe}     Processo Administrativo {processo}", fonte_cnpj)
            ]
            for range_str, texto, fnt in cabecalhos:
                start_cell = range_str.split(":")[0]
                try:
                    ws.merge_cells(range_str)
                    ws[start_cell].value = texto
                    ws[start_cell].font = fnt
                    ws[start_cell].alignment = alinhamento_centro
                    aplicar_borda(range_str)
                except Exception:
                    pass  # evita crash caso a planilha seja pequena

            # Nome do posto
            try:
                ws.merge_cells("A10:C10")
                ws["A10"].value = aba
                ws["A10"].font = fonte_titulo
                ws["A10"].alignment = alinhamento_centro
                aplicar_borda("A10:C10")
            except Exception:
                pass

            # Mês de referência
            try:
                ws.merge_cells("A11:C11")
                ws["A11"].value = f"Mês de Referência: 01 a 31 de {mes} de {ano}"
                ws["A11"].font = fonte_negrito
                ws["A11"].alignment = alinhamento_centro
                aplicar_borda("A11:C11")
            except Exception:
                pass

            # Se a planilha possui o formato esperado (header na linha 15)
            if ws.max_row >= 15:
                # Cabeçalho da tabela (linha 15) -> CAPS LOCK + negrito
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=15, column=col)
                    if cell.value:
                        cell.value = str(cell.value).upper()
                    cell.font = fonte_negrito
                    cell.alignment = alinhamento_centro
                    cell.border = borda

                # Dados da tabela (a partir da linha 16)
                for row in ws.iter_rows(min_row=16, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        if cell.value is not None:
                            cell.font = fonte
                            cell.alignment = alinhamento_centro
                            cell.border = borda
            else:
                # Planilha pequena (ex.: Resumo) — aplicar formatação simples para o conteúdo existente
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        if cell.value is not None:
                            cell.font = fonte
                            cell.alignment = alinhamento_centro
                            cell.border = borda

            ultima_linha = ws.max_row + 2
            # --- Rodapé com mês +1 ---
            meses = [
                "janeiro","fevereiro","março","abril","maio","junho",
                "julho","agosto","setembro","outubro","novembro","dezembro"
            ]
            try:
                indice_mes = meses.index(mes)
                mes_rodape = meses[(indice_mes + 1) % 12]
            except ValueError:
                mes_rodape = mes  # caso o mês não esteja na lista, usar o mesmo

            try:
                ws.merge_cells(start_row=ultima_linha, start_column=1, end_row=ultima_linha, end_column=3)
                ws.cell(row=ultima_linha, column=1).value = f"São Paulo, 01 de {mes_rodape} de {ano}"
                ws.cell(row=ultima_linha, column=1).font = fonte_negrito
                ws.cell(row=ultima_linha, column=1).alignment = alinhamento_centro
            except Exception:
                pass

            # Assinaturas
            linha_assinatura = ultima_linha + 3
            try:
                ws.merge_cells(start_row=linha_assinatura, start_column=1, end_row=linha_assinatura, end_column=2)
                ws.cell(row=linha_assinatura, column=1).value = "________________________"
                ws.cell(row=linha_assinatura, column=1).alignment = Alignment(horizontal="left", vertical="center")

                ws.merge_cells(start_row=linha_assinatura, start_column=3, end_row=linha_assinatura, end_column=4)
                ws.cell(row=linha_assinatura, column=3).value = "________________________"
                ws.cell(row=linha_assinatura, column=3).alignment = alinhamento_centro

                ws.cell(row=linha_assinatura + 1, column=1).value = "Fiscal de Contrato"
                ws.cell(row=linha_assinatura + 1, column=1).font = fonte_negrito
                ws.cell(row=linha_assinatura + 1, column=1).alignment = Alignment(horizontal="left", vertical="center")

                ws.cell(row=linha_assinatura + 1, column=3).value = "Supervisão"
                ws.cell(row=linha_assinatura + 1, column=3).font = fonte_negrito
                ws.cell(row=linha_assinatura + 1, column=3).alignment = alinhamento_centro
            except Exception:
                pass

        wb.save(saida)
        messagebox.showinfo("Sucesso", f"Relatório salvo em:\n{saida}")

    except Exception as e:
        messagebox.showerror("Erro", str(e))


# --- INTERFACE TKINTER ---
root = tk.Tk()
root.title("📊 Gerador de Relatório")
root.geometry("500x350")
root.configure(bg="#f4f4f4")

style = ttk.Style()
style.theme_use("clam")
style.configure("TButton", font=("Segoe UI", 12, "bold"), padding=10,
                background="#4CAF50", foreground="white")
style.map("TButton", background=[("active", "#45a049")])

frame = ttk.Frame(root, padding=20)
frame.pack(expand=True)

titulo = tk.Label(frame, text="Gerador de Relatório por Posto",
                  font=("Segoe UI", 16, "bold"), bg="#f4f4f4", fg="#333")
titulo.pack(pady=10)

frame_data = ttk.Frame(frame)
frame_data.pack(pady=5)

tk.Label(frame_data, text="Mês:", font=("Segoe UI", 11), background="#f4f4f4").grid(row=0, column=0, padx=5)
combo_mes = ttk.Combobox(frame_data, values=[
    "janeiro","fevereiro","março","abril","maio","junho",
    "julho","agosto","setembro","outubro","novembro","dezembro"
], width=12)
combo_mes.grid(row=0, column=1, padx=5)

tk.Label(frame_data, text="Ano:", font=("Segoe UI", 11), background="#f4f4f4").grid(row=0, column=2, padx=5)
entry_ano = tk.Entry(frame_data, width=8)
entry_ano.grid(row=0, column=3, padx=5)

# --- CAMPOS CONTRATO / P.E. / PROCESSO --- 
frame_contrato = ttk.Frame(frame)
frame_contrato.pack(pady=5)

tk.Label(frame_contrato, text="Contrato:", font=("Segoe UI", 11), background="#f4f4f4").grid(row=0, column=0, padx=5)
entry_contrato = tk.Entry(frame_contrato, width=20)
entry_contrato.grid(row=0, column=1, padx=5)

tk.Label(frame_contrato, text="P.E.:", font=("Segoe UI", 11), background="#f4f4f4").grid(row=1, column=0, padx=5)
entry_pe = tk.Entry(frame_contrato, width=20)
entry_pe.grid(row=1, column=1, padx=5)

tk.Label(frame_contrato, text="Processo Adm.:", font=("Segoe UI", 11), background="#f4f4f4").grid(row=2, column=0, padx=5)
entry_processo = tk.Entry(frame_contrato, width=25)
entry_processo.grid(row=2, column=1, padx=5, columnspan=3)

btn = ttk.Button(frame, text="📂 Selecionar e Gerar Relatório", command=gerar_relatorio)
btn.pack(pady=20)

root.mainloop()
